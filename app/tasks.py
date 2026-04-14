import io
import json
import logging
import os
import re
import time
from datetime import datetime, timezone
from decimal import Decimal
from uuid import uuid4

import boto3
import pypdf
from botocore.exceptions import BotoCoreError, ClientError

from app.categorisation import apply_rules, UNRESOLVED_CATEGORIES
from app.celery_app import celery_app
from app.database import SessionLocal
from app.models import Account, AiReport, Case, CaseException, Document, MerchantAlias, ProcessingJob, RiskFlag, Transaction, ValidationResult
from app.risk_flags import compute_risk_flags

logger = logging.getLogger(__name__)

ALLOWED_MIME_TYPES = {
    "application/pdf",
    "text/csv",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "image/jpeg",
    "image/png",
    "image/tiff",
}

ALLOWED_EXTENSIONS = {".pdf", ".csv", ".xls", ".xlsx", ".jpg", ".jpeg", ".png", ".tif", ".tiff"}

SUPPORTED_SOURCE_TYPES = {"upload", "email", "api", "sftp", "manual"}

KNOWN_BANKS = [
    "barclays", "hsbc", "lloyds", "natwest", "santander", "halifax",
    "nationwide", "monzo", "starling", "revolut", "metro bank", "first direct",
    "bank of scotland", "rbs", "royal bank of scotland", "co-operative bank",
    "coop bank", "virgin money", "tsb", "ulster bank", "danske bank",
    "allied irish", "bank of ireland",
]

DATE_PATTERN = re.compile(
    r"\b(\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4}|\d{4}[\/\-\.]\d{1,2}[\/\-\.]\d{1,2}|"
    r"(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\.?\s+\d{1,2},?\s+\d{4}|"
    r"\d{1,2}\s+(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\.?\s+\d{4})\b",
    re.IGNORECASE,
)

# ---------------------------------------------------------------------------
# Transaction-type classification helpers
# ---------------------------------------------------------------------------

# Ordered list of (compiled_pattern, transaction_type_value) pairs.
# Patterns are checked against description_raw; first match wins.
_TXN_TYPE_RULES = [
    (re.compile(r"^\s*BGC\b", re.I), "bank_giro_credit"),
    (re.compile(r"^\s*BACS\b", re.I), "bacs"),
    (re.compile(r"^\s*SO\b", re.I), "standing_order"),
    (re.compile(r"^\s*DD\b", re.I), "direct_debit"),
    (re.compile(r"^\s*(?:FPS|FPO|FPI|FP)\b", re.I), "faster_payment"),
    (re.compile(r"^\s*(?:TFR|TRANSFER)\b", re.I), "transfer"),
    (re.compile(r"^\s*ATM\b", re.I), "atm"),
    (re.compile(r"^\s*BP\b", re.I), "bill_payment"),
    (re.compile(r"^\s*(?:CHQ|CHEQUE)\b", re.I), "cheque"),
    (re.compile(r"^\s*(?:VIS|VISA|DEB|DEBIT CARD|POS)\b", re.I), "card_payment"),
    (re.compile(r"^\s*INT\b", re.I), "interest"),
    (re.compile(r"^\s*(?:REF|REFUND)\b", re.I), "refund"),
    (re.compile(r"\b(?:SALARY|PAYROLL|WAGES)\b", re.I), "income"),
    (re.compile(r"^\s*(?:RECEIVED\s+FROM|PAYMENT\s+FROM)\b", re.I), "credit"),
    (re.compile(r"^\s*(?:CDT|CREDIT)\b", re.I), "credit"),
]

# Counterparty extraction: patterns applied in order; group(1) is the counterparty name.
_COUNTERPARTY_PATTERNS = [
    re.compile(r"^\s*(?:RECEIVED\s+FROM|FROM|PAYMENT\s+FROM)\s+(.+)", re.I),
    re.compile(r"^\s*(?:PAYMENT\s+TO|TRANSFER\s+TO)\s+(.+)", re.I),
    re.compile(r"^\s*BGC\s+(.+)", re.I),
    re.compile(r"^\s*(?:FPS|FPI|FPO)\s+(.+)", re.I),
    re.compile(r"^\s*DD\s+(.+)", re.I),
    re.compile(r"^\s*SO\s+(.+)", re.I),
    re.compile(r"^\s*BACS\s+(.+)", re.I),
]

# Reference-style suffixes to strip from extracted counterparty names
_COUNTERPARTY_SUFFIX = re.compile(r"\s+(?:REF|REFERENCE)\s+\S+$|\s+\d{5,}$", re.I)


def _classify_transaction_type(description_raw: str, direction: str) -> str:
    """Return a transaction type code derived from the description prefix.

    Falls back to ``"other"`` when no known code is detected.
    """
    if not description_raw:
        return "other"
    for pattern, txn_type in _TXN_TYPE_RULES:
        if pattern.search(description_raw):
            return txn_type
    return "other"


def _extract_counterparty_from_description(description_raw: str) -> "str | None":
    """Best-effort extraction of a counterparty name from a raw transaction description.

    Returns ``None`` when no known prefix pattern is matched.
    """
    if not description_raw:
        return None
    for pattern in _COUNTERPARTY_PATTERNS:
        m = pattern.match(description_raw)
        if m:
            name = _COUNTERPARTY_SUFFIX.sub("", m.group(1)).strip()
            if name:
                return name
    return None


def _mark_job_started(db, job_id: str):
    job = db.query(ProcessingJob).filter(ProcessingJob.job_id == job_id).first()
    if job:
        job.status = "running"
        job.started_at = datetime.now(timezone.utc)
        db.commit()
    return job


def _mark_job_completed(db, job_id: str, result: dict):
    job = db.query(ProcessingJob).filter(ProcessingJob.job_id == job_id).first()
    if job:
        job.status = "completed"
        job.completed_at = datetime.now(timezone.utc)
        job.result_json = json.dumps(result)
        db.commit()


def _mark_job_failed(db, job_id: str, error_code: str, error_message: str):
    job = db.query(ProcessingJob).filter(ProcessingJob.job_id == job_id).first()
    if job:
        job.status = "failed"
        job.completed_at = datetime.now(timezone.utc)
        job.error_code = error_code
        job.error_message = error_message
        db.commit()


def _save_validation_result(db, document_id: str, job_id: str, check_name: str,
                             passed: bool, severity: str, result_code: str,
                             message: str, details: dict = None):
    vr = ValidationResult(
        validation_result_id=f"vr_{uuid4().hex[:12]}",
        document_id=document_id,
        job_id=job_id,
        check_name=check_name,
        severity=severity,
        passed=passed,
        result_code=result_code,
        message=message,
        details_json=json.dumps(details) if details else None,
    )
    db.add(vr)
    db.commit()
    return vr


def _save_exception(db, document: Document, job_id: str, exception_type: str,
                    severity: str, title: str, description: str):
    exc = CaseException(
        exception_id=f"exc_{uuid4().hex[:12]}",
        case_id=document.case_id,
        document_id=document.document_id,
        job_id=job_id,
        exception_type=exception_type,
        severity=severity,
        title=title,
        description=description,
    )
    db.add(exc)
    db.commit()


def _fetch_file_from_s3(storage_key: str) -> bytes | None:
    """Fetch raw file bytes from R2/S3. Returns None on any error."""
    bucket = os.getenv("R2_BUCKET_NAME")
    endpoint_url = os.getenv("R2_ENDPOINT_URL")
    if not bucket or not storage_key:
        return None
    try:
        client = boto3.client(
            "s3",
            endpoint_url=endpoint_url,
            region_name=os.getenv("R2_REGION", "auto"),
            aws_access_key_id=os.getenv("R2_ACCESS_KEY_ID"),
            aws_secret_access_key=os.getenv("R2_SECRET_ACCESS_KEY"),
        )
        response = client.get_object(Bucket=bucket, Key=storage_key)
        return response["Body"].read()
    except (BotoCoreError, ClientError) as exc:
        logger.warning("R2 fetch failed for key %s: %s", storage_key, exc)
        return None


def _run_checks(db, document: Document, job_id: str) -> list[dict]:
    """Run all validation checks and persist results. Returns list of check outcomes."""
    outcomes = []
    file_bytes: bytes | None = None

    # --- 1. document_metadata_valid ---
    has_filename = bool(document.original_filename and document.original_filename.strip())
    has_source_type = bool(document.source_type and document.source_type.strip())
    metadata_valid = has_filename and has_source_type
    missing_fields = []
    if not has_filename:
        missing_fields.append("original_filename")
    if not has_source_type:
        missing_fields.append("source_type")
    _save_validation_result(
        db, document.document_id, job_id,
        check_name="document_metadata_valid",
        passed=metadata_valid,
        severity="High",
        result_code="METADATA_VALID" if metadata_valid else "METADATA_INCOMPLETE",
        message="Required document metadata is present" if metadata_valid else f"Missing required fields: {', '.join(missing_fields)}",
        details={"original_filename": document.original_filename, "source_type": document.source_type, "missing_fields": missing_fields},
    )
    if not metadata_valid:
        _save_exception(
            db, document, job_id,
            exception_type="validation_failure",
            severity="High",
            title="Document metadata incomplete",
            description=f"The following required fields are missing or blank: {', '.join(missing_fields)}.",
        )
    outcomes.append({"check": "document_metadata_valid", "passed": metadata_valid})

    # --- 2. source_type_supported ---
    source_type = (document.source_type or "").strip().lower()
    source_supported = source_type in SUPPORTED_SOURCE_TYPES
    _save_validation_result(
        db, document.document_id, job_id,
        check_name="source_type_supported",
        passed=source_supported,
        severity="Medium",
        result_code="SOURCE_TYPE_SUPPORTED" if source_supported else "SOURCE_TYPE_UNSUPPORTED",
        message=f"Source type '{source_type}' is supported" if source_supported else f"Source type '{source_type}' is not in the supported list",
        details={"source_type": source_type, "supported_types": sorted(SUPPORTED_SOURCE_TYPES)},
    )
    if not source_supported:
        _save_exception(
            db, document, job_id,
            exception_type="validation_failure",
            severity="Medium",
            title="Source type not supported",
            description=f"Document source type '{source_type}' is not recognised. Supported types: {', '.join(sorted(SUPPORTED_SOURCE_TYPES))}.",
        )
    outcomes.append({"check": "source_type_supported", "passed": source_supported})

    # --- 3. file_exists ---
    exists = bool(document.storage_key or document.file_size)
    _save_validation_result(
        db, document.document_id, job_id,
        check_name="file_exists",
        passed=exists,
        severity="Critical",
        result_code="FILE_EXISTS" if exists else "FILE_MISSING",
        message="Document record has storage reference" if exists else "No storage key or file size recorded",
        details={"storage_key": document.storage_key, "file_size": document.file_size},
    )
    if not exists:
        _save_exception(
            db, document, job_id,
            exception_type="validation_failure",
            severity="Critical",
            title="Document file reference missing",
            description="The document record has neither a storage key nor a recorded file size.",
        )
    outcomes.append({"check": "file_exists", "passed": exists})

    # --- 4. file_type_allowed ---
    ext = os.path.splitext(document.original_filename or "")[-1].lower()
    mime_ok = document.mime_type in ALLOWED_MIME_TYPES if document.mime_type else False
    ext_ok = ext in ALLOWED_EXTENSIONS
    type_allowed = mime_ok or ext_ok
    _save_validation_result(
        db, document.document_id, job_id,
        check_name="file_type_allowed",
        passed=type_allowed,
        severity="High",
        result_code="TYPE_ALLOWED" if type_allowed else "TYPE_NOT_ALLOWED",
        message=f"MIME type '{document.mime_type}', extension '{ext}'" + ("" if type_allowed else " — not in allowed list"),
        details={"mime_type": document.mime_type, "extension": ext},
    )
    if not type_allowed:
        _save_exception(
            db, document, job_id,
            exception_type="validation_failure",
            severity="High",
            title="File type not allowed",
            description=f"MIME type '{document.mime_type}' and extension '{ext}' are not in the allowed list.",
        )
    outcomes.append({"check": "file_type_allowed", "passed": type_allowed})

    # --- 5. file_not_empty ---
    not_empty = bool(document.file_size and document.file_size > 0)
    _save_validation_result(
        db, document.document_id, job_id,
        check_name="file_not_empty",
        passed=not_empty,
        severity="High",
        result_code="FILE_NOT_EMPTY" if not_empty else "FILE_EMPTY",
        message=f"File size: {document.file_size} bytes" if document.file_size is not None else "File size unknown",
        details={"file_size": document.file_size},
    )
    if not not_empty:
        _save_exception(
            db, document, job_id,
            exception_type="validation_failure",
            severity="High",
            title="File is empty or size unknown",
            description=f"Recorded file size is {document.file_size!r}.",
        )
    outcomes.append({"check": "file_not_empty", "passed": not_empty})

    # --- 6. file_readable (fetch from R2) ---
    if document.storage_key:
        file_bytes = _fetch_file_from_s3(document.storage_key)
        readable = file_bytes is not None
        _save_validation_result(
            db, document.document_id, job_id,
            check_name="file_readable",
            passed=readable,
            severity="Critical",
            result_code="FILE_READABLE" if readable else "FILE_UNREADABLE",
            message="File successfully retrieved from storage" if readable else "Could not retrieve file from storage",
            details={"storage_key": document.storage_key, "bytes_read": len(file_bytes) if file_bytes else 0},
        )
        if not readable:
            _save_exception(
                db, document, job_id,
                exception_type="validation_failure",
                severity="Critical",
                title="File could not be read from storage",
                description=f"Attempt to read '{document.storage_key}' from R2 failed.",
            )
        outcomes.append({"check": "file_readable", "passed": readable})
    else:
        _save_validation_result(
            db, document.document_id, job_id,
            check_name="file_readable",
            passed=False,
            severity="Critical",
            result_code="NO_STORAGE_KEY",
            message="No storage key — cannot attempt read",
        )
        outcomes.append({"check": "file_readable", "passed": False})

    # --- 7. page_count_detected (PDFs only) ---
    is_pdf = (document.mime_type == "application/pdf") or ext == ".pdf"
    if is_pdf and file_bytes:
        try:
            reader = pypdf.PdfReader(io.BytesIO(file_bytes))
            page_count = len(reader.pages)
            detected = page_count > 0
            _save_validation_result(
                db, document.document_id, job_id,
                check_name="page_count_detected",
                passed=detected,
                severity="Medium",
                result_code="PAGE_COUNT_DETECTED" if detected else "NO_PAGES",
                message=f"{page_count} page(s) detected" if detected else "PDF appears to have no pages",
                details={"page_count": page_count},
            )
            if not detected:
                _save_exception(
                    db, document, job_id,
                    exception_type="validation_failure",
                    severity="Medium",
                    title="PDF has no pages",
                    description="The PDF was opened successfully but reports zero pages.",
                )
            outcomes.append({"check": "page_count_detected", "passed": detected, "page_count": page_count})
        except Exception as pdf_err:
            _save_validation_result(
                db, document.document_id, job_id,
                check_name="page_count_detected",
                passed=False,
                severity="Medium",
                result_code="PDF_PARSE_ERROR",
                message=f"Could not parse PDF: {pdf_err}",
            )
            _save_exception(
                db, document, job_id,
                exception_type="validation_failure",
                severity="Medium",
                title="PDF could not be parsed",
                description=str(pdf_err),
            )
            outcomes.append({"check": "page_count_detected", "passed": False})
    else:
        _save_validation_result(
            db, document.document_id, job_id,
            check_name="page_count_detected",
            passed=True,
            severity="Low",
            result_code="NOT_APPLICABLE",
            message="Page count check skipped for non-PDF files",
        )
        outcomes.append({"check": "page_count_detected", "passed": True, "skipped": True})

    # --- 8. statement_date_range_present ---
    extracted_text = ""
    if is_pdf and file_bytes:
        try:
            reader = pypdf.PdfReader(io.BytesIO(file_bytes))
            extracted_text = " ".join(
                (page.extract_text() or "") for page in reader.pages[:5]
            )
        except Exception:
            pass

    dates_found = DATE_PATTERN.findall(extracted_text)
    date_range_present = len(dates_found) >= 2
    _save_validation_result(
        db, document.document_id, job_id,
        check_name="statement_date_range_present",
        passed=date_range_present if extracted_text else True,
        severity="Low",
        result_code="DATE_RANGE_PRESENT" if date_range_present else ("DATE_RANGE_NOT_FOUND" if extracted_text else "TEXT_NOT_EXTRACTED"),
        message=(
            f"Found {len(dates_found)} date(s) in document text" if extracted_text
            else "Date range check skipped — text could not be extracted"
        ),
        details={"dates_found": dates_found[:10]},
    )
    outcomes.append({"check": "statement_date_range_present", "passed": date_range_present if extracted_text else True})

    # --- 9. bank_type_identified ---
    lower_text = extracted_text.lower() + " " + (document.original_filename or "").lower()
    matched_banks = [b for b in KNOWN_BANKS if b in lower_text]
    bank_identified = len(matched_banks) > 0
    _save_validation_result(
        db, document.document_id, job_id,
        check_name="bank_type_identified",
        passed=bank_identified if extracted_text else True,
        severity="Low",
        result_code="BANK_IDENTIFIED" if bank_identified else ("BANK_NOT_IDENTIFIED" if extracted_text else "TEXT_NOT_EXTRACTED"),
        message=(
            f"Bank(s) identified: {', '.join(matched_banks)}" if bank_identified
            else ("No known bank name found in document" if extracted_text else "Bank check skipped — text could not be extracted")
        ),
        details={"matched_banks": matched_banks},
    )
    outcomes.append({"check": "bank_type_identified", "passed": bank_identified if extracted_text else True})

    return outcomes


def _mark_job_completed_with_warnings(db, job_id: str, result: dict):
    job = db.query(ProcessingJob).filter(ProcessingJob.job_id == job_id).first()
    if job:
        job.status = "completed"
        job.completed_at = datetime.now(timezone.utc)
        job.result_json = json.dumps(result)
        db.commit()


def _persist_reconciliation_exceptions(db, reconciliation_result, case_id: str, document_id: str, job_id: str):
    """Create CaseException rows for every finding in *reconciliation_result*."""
    severity_map = {
        "Critical": "Critical",
        "High": "High",
        "Medium": "Medium",
        "Low": "Low",
    }
    for finding in reconciliation_result.findings:
        exception_id = f"exc_{uuid4().hex[:16]}"
        exc = CaseException(
            exception_id=exception_id,
            case_id=case_id,
            document_id=document_id,
            transaction_id=finding.transaction_id,
            job_id=job_id,
            exception_type=finding.check,
            severity=severity_map.get(finding.severity, "Medium"),
            status="Open",
            title=finding.title,
            description=finding.description,
        )
        db.add(exc)
    db.commit()
@celery_app.task(bind=True, name="validate_document_task")
def validate_document_task(self, document_id: str, job_id: str):
    """Async task to validate a document with real file checks."""
    db = SessionLocal()
    try:
        document = db.query(Document).filter(Document.document_id == document_id).first()

        if not document:
            _mark_job_failed(db, job_id, "NOT_FOUND", "Document not found")
            raise ValueError(f"Document {document_id} not found")

        _mark_job_started(db, job_id)

        outcomes = _run_checks(db, document, job_id)

        all_critical_passed = all(
            o["passed"] for o in outcomes
            if o.get("check") in ("file_exists", "file_readable")
        )

        if all_critical_passed:
            document.status = "Validated"
        else:
            document.status = "ValidationFailed"
        db.commit()

        result = {
            "document_id": document_id,
            "status": document.status,
            "checks_run": len(outcomes),
            "checks_passed": sum(1 for o in outcomes if o["passed"]),
            "checks_failed": sum(1 for o in outcomes if not o["passed"]),
            "outcomes": outcomes,
        }
        if all_critical_passed:
            _mark_job_completed(db, job_id, result)
        else:
            _mark_job_failed(db, job_id, "VALIDATION_FAILED", "One or more critical validation checks failed")
        return result

    except Exception as exc:
        _mark_job_failed(db, job_id, "VALIDATION_ERROR", str(exc))
        raise
    finally:
        db.close()


@celery_app.task(bind=True, name="extract_document_task")
def extract_document_task(self, document_id: str, job_id: str):
    """
    Extraction worker that:
      1. Fetches the stored file from S3
      2. Sends it to Google Document AI
      3. Normalises the response
      4. Saves Account and Transaction records
      5. Creates exceptions for low-confidence rows
      6. Runs reconciliation on the extracted data
      7. Updates the document and job status
    """
    from app.documentai import CONFIDENCE_THRESHOLD, process_document
    from app.reconciliation import run_reconciliation
    from app.storage import download_file_from_s3
    db = SessionLocal()
    try:
        document = db.query(Document).filter(Document.document_id == document_id).first()
        if not document:
            _mark_job_failed(db, job_id, "NOT_FOUND", "Document not found")
            raise ValueError(f"Document {document_id} not found")

        if not document.storage_key:
            _mark_job_failed(db, job_id, "NO_STORAGE_KEY", "Document has no storage key")
            raise ValueError(f"Document {document_id} has no storage key")

        _mark_job_started(db, job_id)

        # 1. Download file
        logger.info("Downloading document %s from storage", document_id)
        file_bytes = download_file_from_s3(document.storage_key)
        mime_type = document.mime_type or "application/pdf"

        # 2. Send to Document AI
        logger.info("Sending document %s to Document AI", document_id)
        extraction = process_document(file_bytes, mime_type)

        # Fallback: if Document AI returned no transactions for a PDF, attempt
        # text-based extraction so Barclays (and similar) column layouts are handled.
        if not extraction.transactions and mime_type == "application/pdf":
            logger.warning(
                "Document AI returned 0 transactions for document %s; "
                "attempting PDF text fallback parser.",
                document_id,
            )
            from app.pdf_fallback import extract_from_pdf_text
            fallback = extract_from_pdf_text(file_bytes)
            if fallback.transactions:
                logger.info(
                    "PDF text fallback extracted %d transaction(s) for document %s",
                    len(fallback.transactions), document_id,
                )
                extraction = fallback
            else:
                logger.warning(
                    "PDF text fallback also returned 0 transactions for document %s",
                    document_id,
                )

        # 3. Persist Account
        account_id = f"acc_{uuid4().hex[:8]}"
        acct_data = extraction.account
        account = Account(
            account_id=account_id,
            case_id=document.case_id,
            document_id=document_id,
            bank_name=acct_data.bank_name,
            account_holder_name=acct_data.account_holder_name,
            sort_code=acct_data.sort_code,
            account_number_masked=acct_data.account_number_masked,
            statement_start_date=acct_data.statement_start_date,
            statement_end_date=acct_data.statement_end_date,
            opening_balance=acct_data.opening_balance,
            closing_balance=acct_data.closing_balance,
        )
        db.add(account)
        db.flush()

        # 4. Persist Transactions + create exceptions for uncertain rows
        saved_count = 0
        exception_count = 0
        txn_list = []
        for idx, txn_data in enumerate(extraction.transactions):
            transaction_id = f"txn_{uuid4().hex[:8]}"

            # Classify transaction type from description when the extractor didn't supply one.
            txn_type = _classify_transaction_type(
                txn_data.description_raw or "", txn_data.direction or ""
            )

            # Derive counterparty name from description when the extractor didn't supply one.
            counterparty = txn_data.counterparty_name or _extract_counterparty_from_description(
                txn_data.description_raw
            )

            txn = Transaction(
                transaction_id=transaction_id,
                case_id=document.case_id,
                document_id=document_id,
                account_id=account_id,
                transaction_date=txn_data.transaction_date,
                posting_date=txn_data.posting_date,
                description_raw=txn_data.description_raw,
                description_normalised=txn_data.description_normalised,
                direction=txn_data.direction,
                amount=txn_data.amount,
                balance=txn_data.balance,
                merchant_name=txn_data.merchant_name,
                counterparty_name=counterparty,
                transaction_type=txn_type,
                extractor_confidence=txn_data.extractor_confidence,
                source_page_number=txn_data.source_page_number,
                source_row_reference=str(idx + 1),
            )
            db.add(txn)
            db.flush()
            saved_count += 1
            _amount = float(txn_data.amount) if txn_data.amount is not None else None
            txn_list.append({
                "transaction_id": transaction_id,
                "date": txn_data.transaction_date.isoformat() if txn_data.transaction_date else None,
                "description": txn_data.description_raw,
                "debit": _amount if txn_data.direction == "debit" else None,
                "credit": _amount if txn_data.direction == "credit" else None,
                "balance": float(txn_data.balance) if txn_data.balance is not None else None,
            })

            if txn_data.extractor_confidence < CONFIDENCE_THRESHOLD:
                exception_id = f"exc_{uuid4().hex[:8]}"
                exc_record = CaseException(
                    exception_id=exception_id,
                    case_id=document.case_id,
                    document_id=document_id,
                    transaction_id=transaction_id,
                    job_id=job_id,
                    exception_type="low_confidence_extraction",
                    severity="Medium",
                    status="Open",
                    title="Low-confidence transaction extraction",
                    description=(
                        f"Transaction row {idx + 1} was extracted with confidence "
                        f"{float(txn_data.extractor_confidence):.2f}, below the threshold "
                        f"{float(CONFIDENCE_THRESHOLD):.2f}. Manual review recommended."
                    ),
                )
                db.add(exc_record)
                exception_count += 1

        db.commit()

        # 5. Reconcile the extracted data
        extracted_data = {
            "document_id": document_id,
            "opening_balance": float(acct_data.opening_balance) if acct_data.opening_balance is not None else None,
            "closing_balance": float(acct_data.closing_balance) if acct_data.closing_balance is not None else None,
            "transactions": txn_list,
        }
        reconciliation_result = run_reconciliation(extracted_data)

        _persist_reconciliation_exceptions(
            db,
            reconciliation_result,
            case_id=document.case_id,
            document_id=document_id,
            job_id=job_id,
        )

        outcome = reconciliation_result.outcome  # "passed" | "warning" | "failed"
        finding_summaries = [
            {
                "check": f.check,
                "severity": f.severity,
                "title": f.title,
                "transaction_id": f.transaction_id,
            }
            for f in reconciliation_result.findings
        ]

        result = {
            "document_id": document_id,
            "account_id": account_id,
            "transactions_saved": saved_count,
            "exceptions_created": exception_count,
            "reconciliation": {
                "outcome": outcome,
                "finding_count": len(reconciliation_result.findings),
                "findings": finding_summaries,
            },
        }

        if outcome == "failed":
            document.status = "ExtractionFailed"
            db.commit()
            _mark_job_failed(
                db, job_id, "RECONCILIATION_FAILED",
                f"Reconciliation failed with {len(reconciliation_result.findings)} finding(s).",
            )
            result["status"] = "ExtractionFailed"
            result["message"] = "Extraction reconciliation failed — see exceptions for details."
        elif outcome == "warning":
            document.status = "ExtractionWarning"
            db.commit()
            _mark_job_completed_with_warnings(db, job_id, result)
            result["status"] = "ExtractionWarning"
            result["message"] = "Extraction completed with reconciliation warnings — see exceptions for details."
        else:
            document.status = "Extracted"
            db.commit()
            _mark_job_completed(db, job_id, result)
            result["status"] = "Extracted"
            result["message"] = "Extraction and reconciliation completed successfully."

        logger.info(
            "Extraction complete for %s: %d transactions, %d exceptions, reconciliation=%s",
            document_id, saved_count, exception_count, outcome,
        )
        return result

    except Exception as exc:
        db.rollback()
        _mark_job_failed(db, job_id, "EXTRACTION_ERROR", str(exc))
        logger.exception("Extraction failed for document %s", document_id)
        raise
    finally:
        db.close()


@celery_app.task(bind=True, name="categorise_document_task")
def categorise_document_task(self, document_id: str, job_id: str):
    """Async task to categorise transactions extracted from a document."""
    db = SessionLocal()
    try:
        document = db.query(Document).filter(Document.document_id == document_id).first()

        if not document:
            _mark_job_failed(db, job_id, "NOT_FOUND", "Document not found")
            raise ValueError(f"Document {document_id} not found")

        EXTRACTION_COMPLETE_STATUSES = {"Extracted", "ExtractionWarning", "Categorised"}
        if document.status not in EXTRACTION_COMPLETE_STATUSES:
            _mark_job_failed(
                db, job_id, "EXTRACTION_REQUIRED",
                f"Document status is '{document.status}'. Categorisation requires extraction to complete first "
                f"(expected one of: {', '.join(sorted(EXTRACTION_COMPLETE_STATUSES))}).",
            )
            raise ValueError(
                f"Document {document_id} has not been extracted yet (status: {document.status})"
            )

        _mark_job_started(db, job_id)

        transactions = (
            db.query(Transaction)
            .filter(Transaction.document_id == document_id)
            .all()
        )

        categorised_count = 0
        uncategorised_count = 0
        exceptions_created = 0

        # Load merchant aliases once for the whole batch to avoid N DB queries
        merchant_aliases = db.query(MerchantAlias).all()

        for txn in transactions:
            category, source, rule_id = apply_rules(db, txn, aliases=merchant_aliases)
            txn.category = category
            txn.category_primary = category
            txn.category_source = source
            txn.rule_id = rule_id

            if category in UNRESOLVED_CATEGORIES:
                txn.needs_review = True
                uncategorised_count += 1

                exception_id = f"exc_{uuid4().hex[:8]}"
                description_text = txn.description_raw or txn.description_normalised or "(no description)"
                date_text = str(txn.transaction_date) if txn.transaction_date else "unknown date"
                exc = CaseException(
                    exception_id=exception_id,
                    case_id=document.case_id,
                    document_id=document_id,
                    transaction_id=txn.transaction_id,
                    job_id=job_id,
                    exception_type="UNCATEGORISED_TRANSACTION",
                    severity="Low",
                    status="Open",
                    title="Transaction could not be categorised",
                    description=(
                        f"Transaction '{description_text}' on {date_text} "
                        f"did not match any categorisation rule."
                    ),
                )
                db.add(exc)
                exceptions_created += 1
            else:
                txn.needs_review = False
                categorised_count += 1

        document.status = "Categorised"
        db.commit()

        result = {
            "document_id": document_id,
            "status": "Categorised",
            "transactions_processed": len(transactions),
            "categorised_count": categorised_count,
            "uncategorised_count": uncategorised_count,
            "exceptions_created": exceptions_created,
        }
        _mark_job_completed(db, job_id, result)
        return result
    except Exception as exc:
        _mark_job_failed(db, job_id, "CATEGORISATION_ERROR", str(exc))
        raise
    finally:
        db.close()


@celery_app.task(bind=True, name="generate_report_task")
def generate_report_task(self, case_id: str, report_type: str, job_id: str, report_id: str):
    """Async task to generate an affordability report for a case."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable,
    )

    db = SessionLocal()
    try:
        _mark_job_started(db, job_id)

        # ── 1. Gather case data ──────────────────────────────────────────────
        case = db.query(Case).filter(Case.case_id == case_id).first()
        if not case:
            raise ValueError(f"Case {case_id} not found")

        accounts = db.query(Account).filter(Account.case_id == case_id).all()
        transactions = db.query(Transaction).filter(Transaction.case_id == case_id).all()
        exceptions = db.query(CaseException).filter(CaseException.case_id == case_id).all()

        # ── 2. Summarise transactions by category ────────────────────────────
        category_totals: dict[str, dict] = {}
        for txn in transactions:
            cat = txn.category or "uncategorised"
            if cat not in category_totals:
                category_totals[cat] = {"credits": Decimal("0"), "debits": Decimal("0"), "count": 0}
            amount = txn.amount or Decimal("0")
            # direction field is preferred; fall back to checking credit column when direction is absent
            if txn.direction == "credit" or (not txn.direction and txn.credit and txn.credit > 0):
                category_totals[cat]["credits"] += abs(amount)
            else:
                category_totals[cat]["debits"] += abs(amount)
            category_totals[cat]["count"] += 1

        flagged_transactions = [t for t in transactions if t.needs_review]
        open_exceptions = [e for e in exceptions if e.status == "Open"]

        total_credits = sum(v["credits"] for v in category_totals.values())
        total_debits = sum(v["debits"] for v in category_totals.values())

        # ── 3. Build OpenAI prompt ───────────────────────────────────────────
        category_summary_text = "\n".join(
            f"  - {cat}: {data['count']} transactions, credits £{data['credits']:.2f}, debits £{data['debits']:.2f}"
            for cat, data in sorted(category_totals.items())
        )
        exceptions_text = "\n".join(
            f"  - [{e.severity}] {e.title}: {e.description or ''}"
            for e in open_exceptions[:20]
        ) or "  None"
        flagged_text = f"{len(flagged_transactions)} transactions flagged for review"

        account_text = "\n".join(
            f"  - {a.bank_name or 'Unknown bank'}, account {a.account_number_masked or 'N/A'}, "
            f"opening balance £{a.opening_balance or 0:.2f}, closing balance £{a.closing_balance or 0:.2f}, "
            f"period {a.statement_start_date} to {a.statement_end_date}"
            for a in accounts
        ) or "  No account data available"

        prompt = f"""You are a financial analyst producing an affordability assessment for a mortgage or lending application.

Case reference: {case.case_reference}
Customer: {case.customer_name}
Jurisdiction: {case.jurisdiction}

Bank Accounts:
{account_text}

Transaction Category Summary (total credits £{total_credits:.2f}, total debits £{total_debits:.2f}):
{category_summary_text}

Open Exceptions ({len(open_exceptions)} total):
{exceptions_text}

Flagged: {flagged_text}

Produce a structured affordability report as JSON with exactly these keys:
{{
  "overall_assessment": "<Acceptable|Marginal|Unacceptable>",
  "summary": "<2-3 sentence plain-English summary>",
  "monthly_income_estimate": <number or null>,
  "monthly_expenditure_estimate": <number or null>,
  "disposable_income_estimate": <number or null>,
  "income_sources": ["<source>"],
  "concerns": ["<concern>"],
  "positive_indicators": ["<indicator>"],
  "risk_flags": ["<flag>"],
  "recommendations": ["<recommendation>"]
}}

Respond with valid JSON only, no prose outside the JSON block."""

        # ── 4. Call OpenAI ───────────────────────────────────────────────────
        openai_api_key = os.getenv("OPENAI_API_KEY")
        ai_output: dict = {}
        if openai_api_key:
            from openai import OpenAI
            client = OpenAI(api_key=openai_api_key)
            model = os.getenv("OPENAI_MODEL", "gpt-4o-mini")
            response = client.chat.completions.create(
                model=model,
                messages=[{"role": "user", "content": prompt}],
                temperature=0.2,
                response_format={"type": "json_object"},
            )
            raw = response.choices[0].message.content or "{}"
            ai_output = json.loads(raw)
        else:
            logger.warning("OPENAI_API_KEY not set – using placeholder affordability output")
            # Divide by 3 assuming a standard 3-month bank statement period
            ai_output = {
                "overall_assessment": "Marginal",
                "summary": "OpenAI API key not configured. This is a placeholder assessment.",
                "monthly_income_estimate": float(total_credits / 3) if total_credits else None,
                "monthly_expenditure_estimate": float(total_debits / 3) if total_debits else None,
                "disposable_income_estimate": float((total_credits - total_debits) / 3) if total_credits else None,
                "income_sources": [],
                "concerns": ["OpenAI key not configured – assessment is unverified"],
                "positive_indicators": [],
                "risk_flags": [f"{len(open_exceptions)} open exceptions", f"{len(flagged_transactions)} flagged transactions"],
                "recommendations": ["Configure OPENAI_API_KEY to generate a real assessment"],
            }

        # ── 5. Generate PDF ──────────────────────────────────────────────────
        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=A4,
            leftMargin=2 * cm,
            rightMargin=2 * cm,
            topMargin=2 * cm,
            bottomMargin=2 * cm,
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("Title2", parent=styles["Heading1"], fontSize=16, spaceAfter=6)
        h2_style = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=12, spaceAfter=4)
        body_style = styles["Normal"]
        small_style = ParagraphStyle("Small", parent=styles["Normal"], fontSize=8)

        assessment_colour = {
            "Acceptable": colors.HexColor("#28a745"),
            "Marginal": colors.HexColor("#fd7e14"),
            "Unacceptable": colors.HexColor("#dc3545"),
        }.get(ai_output.get("overall_assessment", ""), colors.grey)

        story = [
            Paragraph("Affordability Report", title_style),
            Paragraph(f"Case Reference: {case.case_reference}", body_style),
            Paragraph(f"Customer: {case.customer_name}", body_style),
            Paragraph(f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}", small_style),
            Spacer(1, 0.4 * cm),
            HRFlowable(width="100%", thickness=1, color=colors.grey),
            Spacer(1, 0.4 * cm),
        ]

        # Overall assessment banner
        banner_data = [[Paragraph(
            f"Overall Assessment: <b>{ai_output.get('overall_assessment', 'N/A')}</b>",
            ParagraphStyle("Banner", parent=styles["Normal"], fontSize=13, textColor=colors.white),
        )]]
        banner_table = Table(banner_data, colWidths=["100%"])
        banner_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), assessment_colour),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("LEFTPADDING", (0, 0), (-1, -1), 12),
        ]))
        story += [banner_table, Spacer(1, 0.5 * cm)]

        story += [Paragraph("Summary", h2_style), Paragraph(ai_output.get("summary", ""), body_style), Spacer(1, 0.4 * cm)]

        # Financials table
        fin_rows = [
            [Paragraph("<b>Metric</b>", body_style), Paragraph("<b>Amount</b>", body_style)],
        ]
        for label, key in [
            ("Est. Monthly Income", "monthly_income_estimate"),
            ("Est. Monthly Expenditure", "monthly_expenditure_estimate"),
            ("Est. Disposable Income", "disposable_income_estimate"),
        ]:
            val = ai_output.get(key)
            fin_rows.append([label, f"£{val:,.2f}" if val is not None else "N/A"])
        fin_table = Table(fin_rows, colWidths=[9 * cm, 7 * cm])
        fin_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#343a40")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#f8f9fa"), colors.white]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dee2e6")),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ]))
        story += [Paragraph("Financial Estimates", h2_style), fin_table, Spacer(1, 0.5 * cm)]

        # Category breakdown table
        cat_rows = [
            [Paragraph("<b>Category</b>", body_style), Paragraph("<b>Txns</b>", body_style),
             Paragraph("<b>Credits</b>", body_style), Paragraph("<b>Debits</b>", body_style)],
        ]
        for cat, data in sorted(category_totals.items()):
            cat_rows.append([cat, str(data["count"]), f"£{data['credits']:,.2f}", f"£{data['debits']:,.2f}"])
        if len(cat_rows) > 1:
            cat_table = Table(cat_rows, colWidths=[7 * cm, 2.5 * cm, 4 * cm, 4 * cm])
            cat_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#343a40")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#f8f9fa"), colors.white]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dee2e6")),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ]))
            story += [Paragraph("Transaction Category Breakdown", h2_style), cat_table, Spacer(1, 0.5 * cm)]

        def _bullet_section(title, items):
            if not items:
                return []
            return [Paragraph(title, h2_style)] + [
                Paragraph(f"• {item}", body_style) for item in items
            ] + [Spacer(1, 0.4 * cm)]

        story += _bullet_section("Income Sources", ai_output.get("income_sources", []))
        story += _bullet_section("Positive Indicators", ai_output.get("positive_indicators", []))
        story += _bullet_section("Concerns", ai_output.get("concerns", []))
        story += _bullet_section("Risk Flags", ai_output.get("risk_flags", []))
        story += _bullet_section("Recommendations", ai_output.get("recommendations", []))

        # Exceptions summary
        if open_exceptions:
            exc_rows = [[Paragraph("<b>Severity</b>", body_style), Paragraph("<b>Title</b>", body_style)]]
            for e in open_exceptions[:50]:
                exc_rows.append([e.severity, e.title])
            exc_table = Table(exc_rows, colWidths=[3 * cm, 14 * cm])
            exc_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#343a40")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#fff3cd"), colors.white]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dee2e6")),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ]))
            story += [Paragraph(f"Open Exceptions ({len(open_exceptions)})", h2_style), exc_table]

        doc.build(story)
        pdf_bytes = pdf_buffer.getvalue()

        # ── 6. Generate XLSX ─────────────────────────────────────────────────
        wb = openpyxl.Workbook()

        # ── Sheet 1: Summary ──
        ws_summary = wb.active
        ws_summary.title = "Summary"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="343A40")
        centre = Alignment(horizontal="center")

        ws_summary.append(["Affordability Report"])
        ws_summary["A1"].font = Font(bold=True, size=14)
        ws_summary.append(["Case Reference", case.case_reference])
        ws_summary.append(["Customer", case.customer_name])
        ws_summary.append(["Jurisdiction", case.jurisdiction])
        ws_summary.append(["Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")])
        ws_summary.append([])
        ws_summary.append(["Overall Assessment", ai_output.get("overall_assessment", "N/A")])
        ws_summary.append(["Summary", ai_output.get("summary", "")])
        ws_summary.append([])
        ws_summary.append(["Est. Monthly Income", ai_output.get("monthly_income_estimate")])
        ws_summary.append(["Est. Monthly Expenditure", ai_output.get("monthly_expenditure_estimate")])
        ws_summary.append(["Est. Disposable Income", ai_output.get("disposable_income_estimate")])
        ws_summary.column_dimensions["A"].width = 30
        ws_summary.column_dimensions["B"].width = 60

        for section_title, key in [
            ("Income Sources", "income_sources"),
            ("Positive Indicators", "positive_indicators"),
            ("Concerns", "concerns"),
            ("Risk Flags", "risk_flags"),
            ("Recommendations", "recommendations"),
        ]:
            ws_summary.append([])
            row = ws_summary.max_row + 1
            ws_summary.append([section_title])
            cell = ws_summary.cell(row=row, column=1)
            cell.font = Font(bold=True)
            for item in ai_output.get(key, []):
                ws_summary.append(["", item])

        # ── Sheet 2: Transactions ──
        ws_txn = wb.create_sheet("Transactions")
        txn_headers = ["transaction_id", "date", "description", "direction", "amount", "balance", "category", "needs_review"]
        ws_txn.append(txn_headers)
        for cell in ws_txn[1]:
            cell.font = header_font
            cell.fill = header_fill
        for txn in transactions:
            ws_txn.append([
                txn.transaction_id,
                str(txn.transaction_date) if txn.transaction_date else "",
                txn.description_normalised or txn.description_raw or "",
                txn.direction or "",
                float(txn.amount) if txn.amount is not None else None,
                float(txn.balance) if txn.balance is not None else None,
                txn.category or "",
                txn.needs_review,
            ])
        for col in ws_txn.columns:
            ws_txn.column_dimensions[col[0].column_letter].width = 20

        # ── Sheet 3: Category Breakdown ──
        ws_cat = wb.create_sheet("Category Breakdown")
        cat_headers = ["Category", "Transaction Count", "Total Credits (£)", "Total Debits (£)"]
        ws_cat.append(cat_headers)
        for cell in ws_cat[1]:
            cell.font = header_font
            cell.fill = header_fill
        for cat, data in sorted(category_totals.items()):
            ws_cat.append([cat, data["count"], float(data["credits"]), float(data["debits"])])
        for col in ws_cat.columns:
            ws_cat.column_dimensions[col[0].column_letter].width = 25

        # ── Sheet 4: Exceptions ──
        ws_exc = wb.create_sheet("Exceptions")
        exc_headers = ["exception_id", "type", "severity", "status", "title", "description"]
        ws_exc.append(exc_headers)
        for cell in ws_exc[1]:
            cell.font = header_font
            cell.fill = header_fill
        for e in exceptions:
            ws_exc.append([e.exception_id, e.exception_type, e.severity, e.status, e.title, e.description or ""])
        for col in ws_exc.columns:
            ws_exc.column_dimensions[col[0].column_letter].width = 25

        xlsx_buffer = io.BytesIO()
        wb.save(xlsx_buffer)
        xlsx_bytes = xlsx_buffer.getvalue()

        # ── 7. Upload to S3 ──────────────────────────────────────────────────
        aws_bucket = os.getenv("AWS_S3_BUCKET")
        pdf_url: str | None = None
        xlsx_url: str | None = None

        if aws_bucket:
            aws_key_id = os.getenv("AWS_ACCESS_KEY_ID")
            aws_secret = os.getenv("AWS_SECRET_ACCESS_KEY")
            aws_region = os.getenv("AWS_REGION", "us-east-1")
            s3 = boto3.client(
                "s3",
                region_name=aws_region,
                aws_access_key_id=aws_key_id,
                aws_secret_access_key=aws_secret,
            )
            pdf_key = f"reports/{case_id}/{report_id}.pdf"
            xlsx_key = f"reports/{case_id}/{report_id}.xlsx"
            s3.put_object(Bucket=aws_bucket, Key=pdf_key, Body=pdf_bytes, ContentType="application/pdf")
            s3.put_object(
                Bucket=aws_bucket, Key=xlsx_key, Body=xlsx_bytes,
                ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            pdf_url = f"https://{aws_bucket}.s3.{aws_region}.amazonaws.com/{pdf_key}"
            xlsx_url = f"https://{aws_bucket}.s3.{aws_region}.amazonaws.com/{xlsx_key}"
        else:
            logger.warning("AWS_S3_BUCKET not configured – report files not persisted to S3")

        # ── 8. Update AiReport record ────────────────────────────────────────
        ai_report = db.query(AiReport).filter(AiReport.report_id == report_id).first()
        if ai_report:
            ai_report.status = "Completed"
            ai_report.completed_at = datetime.now(timezone.utc)
            ai_report.output_json = json.dumps(ai_output)
            ai_report.pdf_file_url = pdf_url
            ai_report.spreadsheet_file_url = xlsx_url
            db.commit()

        result = {
            "case_id": case_id,
            "report_id": report_id,
            "report_type": report_type,
            "status": "Completed",
            "overall_assessment": ai_output.get("overall_assessment"),
            "pdf_file_url": pdf_url,
            "spreadsheet_file_url": xlsx_url,
        }
        _mark_job_completed(db, job_id, result)
        return result

    except Exception as exc:
        # Mark AiReport as Failed
        try:
            ai_report = db.query(AiReport).filter(AiReport.report_id == report_id).first()
            if ai_report:
                ai_report.status = "Failed"
                ai_report.completed_at = datetime.now(timezone.utc)
                db.commit()
        except Exception:
            pass
        _mark_job_failed(db, job_id, "REPORT_ERROR", str(exc))
        raise
    finally:
        db.close()


@celery_app.task(bind=True, name="compute_risk_flags_task")
def compute_risk_flags_task(self, document_id: str, job_id: str):
    """Async task to compute deterministic risk flags for all transactions in a document."""
    db = SessionLocal()
    try:
        document = db.query(Document).filter(Document.document_id == document_id).first()

        if not document:
            _mark_job_failed(db, job_id, "NOT_FOUND", "Document not found")
            raise ValueError(f"Document {document_id} not found")

        _mark_job_started(db, job_id)

        transactions = (
            db.query(Transaction)
            .filter(Transaction.document_id == document_id)
            .all()
        )

        # Remove any previously computed flags for this document so the task is idempotent.
        db.query(RiskFlag).filter(RiskFlag.document_id == document_id).delete()

        flag_dicts = compute_risk_flags(transactions)

        for fd in flag_dicts:
            flag = RiskFlag(
                flag_id=f"rf_{uuid4().hex[:8]}",
                case_id=document.case_id,
                document_id=document_id,
                transaction_id=fd.get("transaction_id"),
                flag_type=fd["flag_type"],
                severity=fd["severity"],
                title=fd["title"],
                detail=fd.get("detail"),
                metric_value=fd.get("metric_value"),
                evidence_summary=fd.get("evidence_summary"),
            )
            db.add(flag)

        db.commit()

        # Build a per-type summary for the job result.
        summary: dict[str, int] = {}
        for fd in flag_dicts:
            summary[fd["flag_type"]] = summary.get(fd["flag_type"], 0) + 1

        result = {
            "document_id": document_id,
            "status": "Completed",
            "flags_created": len(flag_dicts),
            "summary": summary,
        }
        _mark_job_completed(db, job_id, result)
        return result
    except Exception as exc:
        _mark_job_failed(db, job_id, "RISK_FLAGS_ERROR", str(exc))
        raise
    finally:
        db.close()

